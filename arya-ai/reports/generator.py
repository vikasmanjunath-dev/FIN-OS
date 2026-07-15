"""
Arya AI — Report Generator.
Produces self-contained HTML reports by default (no PDF dependency needed) —
or a PDF rendering of the exact same HTML/CSS when fmt="pdf" is passed.

PDF needs WeasyPrint + Pango (a real system library — `brew install pango`,
not just pip). On this machine, WeasyPrint also needs DYLD_LIBRARY_PATH to
include /opt/homebrew/lib at runtime, or it fails with "cannot load library
libgobject-2.0-0" — Homebrew-on-Apple-Silicon doesn't put /opt/homebrew/lib on
the default dynamic-linker search path for a non-Homebrew Python. start.sh
sets this; if you're running uvicorn directly, set it yourself.
"""
import time
import json
from pathlib import Path
from config import REPORTS_DIR


def _write(html: str, base_name: str, fmt: str) -> str:
    if fmt == "pdf":
        from weasyprint import HTML
        fname = REPORTS_DIR / f"{base_name}.pdf"
        HTML(string=html).write_pdf(str(fname))
    else:
        fname = REPORTS_DIR / f"{base_name}.html"
        fname.write_text(html, encoding="utf-8")
    return str(fname)


# ── Portfolio Excel Export ──────────────────────────────────────────────────
# Holdings & unrealized P&L only — NOT a capital-gains tax computation. The
# `holdings` table has no purchase-date column anywhere in this codebase
# (checked: js/finos-context.js, alerts/alert-engine.py — both only select
# symbol/quantity/avg_price/current_price/asset_type). STCG (<12mo) vs LTCG
# (>=12mo) classification needs a lot-level purchase date; faking one would
# produce a wrong tax number, which is worse than not having the feature.

def portfolio_excel(summary: dict) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Holdings"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    green = Font(color="0A7D33")
    red = Font(color="C0392B")

    headers = ["Symbol", "Quantity", "Avg Price", "Current Price", "Price Source",
               "Invested Value", "Current Value", "Unrealized P&L", "P&L %"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for h in summary.get("holdings", []):
        row = [
            h["symbol"], h["quantity"], h["avg_price"], h["current_price"], h["price_source"],
            h["invested_value"], h["current_value"], h["pnl"], h["pnl_pct"],
        ]
        ws.append(row)
        pnl_cell = ws.cell(row=ws.max_row, column=8)
        pnl_cell.font = green if h["pnl"] >= 0 else red

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    # Totals row
    ws.append([])
    total_row = ["TOTAL", "", "", "", "",
                 summary.get("total_invested", 0), summary.get("total_current_value", 0),
                 summary.get("total_pnl", 0), summary.get("total_pnl_pct", 0)]
    ws.append(total_row)
    for col in range(1, len(headers) + 1):
        ws.cell(row=ws.max_row, column=col).font = Font(bold=True)

    # Honest caveat sheet — not buried in fine print
    notes = wb.create_sheet("Notes")
    notes.append(["What this is"])
    notes.append(["Holdings and unrealized profit/loss, with each position's price",
                  "refreshed live at export time (see 'Price Source' column)."])
    notes.append([])
    notes.append(["What this is NOT"])
    notes.append(["A capital-gains tax computation. STCG (<12mo) vs LTCG (>=12mo)",
                  "classification needs a lot-level purchase date, which does not exist",
                  "anywhere in the holdings table this data comes from. Building that",
                  "split without a real date would mean guessing — not done here."])
    notes.column_dimensions["A"].width = 70

    fname = REPORTS_DIR / f"portfolio_{summary.get('user_id','unknown')[:8]}_{int(time.time())}.xlsx"
    wb.save(str(fname))
    return str(fname)


def _ts() -> str:
    return time.strftime("%d %b %Y, %I:%M %p IST")


def _inr(v) -> str:
    if v is None:
        return "N/A"
    v = float(v)
    if v >= 10_000_000:
        return f"₹{v/10_000_000:.2f} Cr"
    if v >= 100_000:
        return f"₹{v/100_000:.2f} L"
    if v >= 1000:
        return f"₹{v:,.0f}"
    return f"₹{v:.2f}"


_BASE_CSS = """
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: -apple-system, 'Helvetica Neue', sans-serif; background: #0a0d18;
       color: #e8eaf0; padding: 32px; line-height: 1.6; }
h1 { font-size: 24px; font-weight: 800; color: #fff; margin-bottom: 4px; }
h2 { font-size: 16px; font-weight: 700; color: #00d4ff; margin: 24px 0 12px;
     padding-bottom: 6px; border-bottom: 1px solid rgba(0,212,255,.2); }
.subtitle { font-size: 11px; color: rgba(255,255,255,.35); margin-bottom: 28px; }
.card { background: rgba(255,255,255,.04); border: 1px solid rgba(255,255,255,.08);
        border-radius: 12px; padding: 16px 20px; margin-bottom: 14px; }
.grid-2 { display: grid; grid-template-columns: 1fr 1fr; gap: 14px; }
.grid-3 { display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 14px; }
.stat-label { font-size: 10px; color: rgba(255,255,255,.4); font-weight: 600;
              text-transform: uppercase; letter-spacing: .4px; margin-bottom: 4px; }
.stat-value { font-size: 18px; font-weight: 800; color: #fff; }
.stat-change { font-size: 11px; margin-top: 2px; }
.green { color: #4dffb4; } .red { color: #ff4d6d; } .gold { color: #ffd93d; }
.cyan  { color: #00d4ff; }
table { width: 100%; border-collapse: collapse; font-size: 12px; margin-top: 8px; }
th { text-align: left; font-size: 10px; color: rgba(255,255,255,.4); font-weight: 700;
     letter-spacing: .4px; padding: 8px 12px; border-bottom: 1px solid rgba(255,255,255,.08); }
td { padding: 9px 12px; border-bottom: 1px solid rgba(255,255,255,.04); }
tr:last-child td { border-bottom: none; }
.badge { display: inline-block; padding: 2px 8px; border-radius: 10px; font-size: 10px;
         font-weight: 700; }
.badge-buy  { background: rgba(77,255,180,.1); color: #4dffb4; }
.badge-sell { background: rgba(255,77,109,.1); color: #ff4d6d; }
.badge-hold { background: rgba(255,211,0,.1);  color: #ffd93d; }
.footer { font-size: 10px; color: rgba(255,255,255,.2); text-align: center;
          margin-top: 40px; padding-top: 16px; border-top: 1px solid rgba(255,255,255,.06); }
"""


def _wrap(title: str, body: str) -> str:
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{title} — Arya AI</title>
<style>{_BASE_CSS}</style>
</head>
<body>
{body}
<div class="footer">Generated by Arya AI · {_ts()} · Data sourced from NSE, BSE, AMFI, CoinGecko</div>
</body>
</html>"""


# ── Quote Report ──────────────────────────────────────────────────────────────

def quote_report(quote: dict, technicals: dict | None = None, fmt: str = "html") -> str:
    sym = quote.get("symbol", "UNKNOWN")
    chg = quote.get("change_pct", 0) or 0
    chg_cls = "green" if chg >= 0 else "red"
    chg_sign = "▲" if chg >= 0 else "▼"
    ta_section = ""
    if technicals and "signals" in technicals:
        verdict = technicals.get("verdict", "HOLD")
        score   = technicals.get("score", 50)
        vbadge  = f"badge-{'buy' if verdict=='BUY' else 'sell' if verdict=='SELL' else 'hold'}"
        rows = "".join(
            f"<tr><td>{k}</td><td>{v.get('signal','—')}</td></tr>"
            for k, v in technicals["signals"].items()
        )
        ta_section = f"""
        <h2>📐 Technical Analysis</h2>
        <div class="card">
          <div style="display:flex;align-items:center;gap:12px;margin-bottom:14px">
            <div><div class="stat-label">Verdict</div>
                 <span class="badge {vbadge}" style="font-size:16px;padding:4px 14px">{verdict}</span></div>
            <div><div class="stat-label">Bullish Score</div>
                 <div class="stat-value {'green' if score>=65 else 'red' if score<=35 else 'gold'}">{score}/100</div></div>
          </div>
          <table><thead><tr><th>Indicator</th><th>Signal</th></tr></thead>
          <tbody>{rows}</tbody></table>
        </div>"""

    body = f"""
    <h1>📊 {sym} Stock Report</h1>
    <div class="subtitle">Exchange: {quote.get('exchange','NSE')} · As of {_ts()}</div>
    <div class="grid-3">
      <div class="card">
        <div class="stat-label">Current Price</div>
        <div class="stat-value">{_inr(quote.get('price'))}</div>
        <div class="stat-change {chg_cls}">{chg_sign} {abs(chg):.2f}%</div>
      </div>
      <div class="card">
        <div class="stat-label">Day High / Low</div>
        <div class="stat-value" style="font-size:14px">{_inr(quote.get('high'))} / {_inr(quote.get('low'))}</div>
      </div>
      <div class="card">
        <div class="stat-label">52W High / Low</div>
        <div class="stat-value" style="font-size:14px">{_inr(quote.get('week52_high'))} / {_inr(quote.get('week52_low'))}</div>
      </div>
      <div class="card">
        <div class="stat-label">Prev Close</div>
        <div class="stat-value">{_inr(quote.get('prev_close'))}</div>
      </div>
      <div class="card">
        <div class="stat-label">Open</div>
        <div class="stat-value">{_inr(quote.get('open'))}</div>
      </div>
      <div class="card">
        <div class="stat-label">Volume</div>
        <div class="stat-value" style="font-size:15px">{quote.get('volume') or 'N/A'}</div>
      </div>
    </div>
    {ta_section}"""

    html = _wrap(f"{sym} Report", body)
    return _write(html, f"{sym}_report_{int(time.time())}", fmt)


# ── Market Overview Report ────────────────────────────────────────────────────

def market_overview_report(overview: dict, commodities: dict,
                            crypto: dict, news: list[dict], fmt: str = "html") -> str:
    # Indices table
    idx_rows = ""
    for name, data in (overview.get("indices") or {}).items():
        chg  = data.get("change_pct", 0) or 0
        cls  = "green" if chg >= 0 else "red"
        sign = "▲" if chg >= 0 else "▼"
        idx_rows += f"""<tr>
          <td><b>{name}</b></td>
          <td>{data.get('price','—')}</td>
          <td class="{cls}">{sign} {abs(chg):.2f}%</td>
        </tr>"""

    # Commodities
    com_rows = ""
    for name, data in (commodities.get("commodities") or {}).items():
        chg  = data.get("change_pct", 0) or 0
        cls  = "green" if chg >= 0 else "red"
        sign = "▲" if chg >= 0 else "▼"
        com_rows += f"""<tr>
          <td><b>{name}</b></td>
          <td>{data.get('price','—')} {data.get('unit','')}</td>
          <td class="{cls}">{sign} {abs(chg):.2f}%</td>
        </tr>"""

    # News
    news_rows = ""
    for art in news[:8]:
        sent = art.get("sentiment", "neutral")
        cls  = "green" if sent == "bullish" else "red" if sent == "bearish" else ""
        news_rows += f"""<tr>
          <td><a href="{art.get('url','')}" target="_blank" style="color:#00d4ff;text-decoration:none">
              {art.get('title','')[:80]}</a></td>
          <td class="{cls}">{sent.upper()}</td>
          <td style="color:rgba(255,255,255,.35)">{art.get('source','')}</td>
        </tr>"""

    body = f"""
    <h1>🌐 Market Overview Report</h1>
    <div class="subtitle">Generated: {_ts()}</div>
    <h2>📈 Indian Indices</h2>
    <div class="card">
      <table><thead><tr><th>Index</th><th>Price</th><th>Change</th></tr></thead>
      <tbody>{idx_rows}</tbody></table>
    </div>
    <div class="grid-2">
      <div>
        <h2>🥇 Commodities</h2>
        <div class="card">
          <table><thead><tr><th>Commodity</th><th>Price</th><th>Change</th></tr></thead>
          <tbody>{com_rows}</tbody></table>
        </div>
      </div>
      <div>
        <h2>🗞️ Top News</h2>
        <div class="card">
          <table><thead><tr><th>Headline</th><th>Sentiment</th><th>Source</th></tr></thead>
          <tbody>{news_rows}</tbody></table>
        </div>
      </div>
    </div>"""

    html = _wrap("Market Overview", body)
    return _write(html, f"market_overview_{int(time.time())}", fmt)
